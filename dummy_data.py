from openpyxl.workbook import Workbook
from faker import Faker

#This program generates dummy test data and write it to an Excel file

wb = Workbook()
ws = wb.active
ws["A1"] = "NAME"
ws["B1"] = "USERNAME"
ws["C1"] = "EMAIL"

fk = Faker()

for i in range(2,21):
    for j in range(1,4):
        ws.cell(row = i, column = 1).value = fk.name()
        ws.cell(row = i, column = 2).value = fk.user_name()
        ws.cell(row = i, column = 3).value = fk.email()

wb.save("dummy_data.xlsx")

#Read the dummy data from Excel file

from openpyxl import load_workbook

wbs = load_workbook("dummy_data.xlsx")
wb1 = wbs.active

# Count the total number of rows

row_ct = wb1.max_row
print(row_ct)

# Count the number of columns

col_ct = wb1.max_column
print(col_ct)

for i in range(2,row_ct+1):
    for j in range(1,col_ct+1):
      print(wb1.cell(row = i, column = j).value, end = ' ')

    print('\n')











